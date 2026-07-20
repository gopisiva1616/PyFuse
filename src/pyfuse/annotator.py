#!/usr/bin/env python3

"""Main module to run the pyfuse fusion annotation pipeline"""
__status__ = "release"
__date__ = "05/05/2020"

import os
import sys
import logging
from pyfuse.utils.common_utils import utils, config
from pyfuse.utils.fus_table2html import write_df2html
from pyfuse.preprocess.preprocess_fusion import PrepareInputBkpt
from pyfuse.annotators import exon_annotations as ex_annot
from pyfuse.annotators import mane_annotation as mane_annot
from pyfuse.annotators import add_getex_cosmic_annotation as cos_gtex
from pyfuse.annotators import fusion_frame_status as frame_calc
from pyfuse.annotators import fusion_sequence as fus_seq
from pyfuse.annotators import add_black_list_genes as bl_list
from pyfuse.annotators.fusion_plot_links import add_fusion_plot_links
from pyfuse.annotators.fusion2vcf import fusion_df_to_bnd_vcf
from openpyxl.utils import get_column_letter

logger = logging.getLogger()


class RunPyfuse():
    """
    A class to annotate input fusion breakpoints

    Attributes
    ----------
    input_bkpt: file-path
        path of input breakpoint file
    out_path: path
        path to write output and log files
    res_path: path
        path of resource files needed for package annotation
    format: str
        fusion caller name to preprocess input breakpoint file according to
        respective caller output format
    target: file-path
        input bed file to filter breakpoint that are within this bed region

    Methods
    -------
    run_pipeline(object):
        handles each method calls i/o to obtains different types of annotation
        from different modules of the package
    preprocess_fusion_bkpt(object):
        preprocess input bkpt file according to format, filters them based on
        different region and return breakpoints in a format required for downstream process
    add_exon_annotation(dataframe):
        adds exon level annotation for each bkpts
    add_cosmic_gtex_annots(dataframe):
        add annotation from cosmic and gtex databases
    add_frame_annotation(dataframe):
        calculated frame status for each breakpoint based on exon annotations
    add_fusion_sequence(dataframe):
        add fusion sequence if reference genome is provided by user
    join_op(object, file-name):
        joins file-name with determined output path
    write_output(object):
        write final annotated DF and generates output files
    """

    def __init__(self, input_bkpt, target_df, format, ref=None, genome=None, fusion_plot_mode="external"):
        self.fusion_bkpt = input_bkpt
        self.target_df = target_df
        self.format = format
        self.ref = ref
        self.genome = self._display_genome(genome)
        self.fusion_plot_mode = fusion_plot_mode

        #print(f'SET OUTPUT PATH in PYFUSE: {utils.out_path}')

    def run_pipeline(self):

        logger.info("Initiating fusion annotation program..")


        # preprocessing input breakpoints
        filtered_bkpts = self.preprocess_fusion_bkpt()
        if filtered_bkpts.empty:
            logger.warning('All input breakpoints were either filtered during \
                preprocessing or no valid breakpoints available for annotation')
            sys.exit(1)
        else:
            logger.info(f'Filtered breakpoints: {filtered_bkpts.shape[0]}')


        # Running exon level annotation
        annotated_df = self.add_exon_annotation(filtered_bkpts)
        annotated_df['Genome'] = self.genome
        # Adding frame calculation

        annotated_df = self.add_frame_annotation(annotated_df)


        # Annotating Cosmic and Gtex
        if 'gtex_fusions' in config or 'cosmic_fusions' in config:
            logger.warning(
                "External fusion resources (GTEx/COSMIC) are enabled. "
                "Confirm redistribution and commercial-use terms for these sources before publishing outputs."
            )
            annotated_df = self.add_cosmic_gtex_annots(annotated_df)

        # Annotating black list fusions
        if 'black_list' in config:
            annotated_df = self.add_black_list_genes(annotated_df)


        # Adding MANE transcript annotation
        if 'mane_status' in config:
            logger.info('MANE status resource found in config; adding MANE annotation')
            annotated_df = self.add_mane_annotation(annotated_df)
        else:
            logger.info('MANE status resource not found in config; skipping MANE annotation and filling columns with \'.\'')

        # Adding fusion seq info
        if self.ref:
            logger.info('reference genome file provided as input')
            annotated_df = self.add_fusion_sequence(annotated_df)
        else:
            logger.info('Skipping sequence annotation - reference not provided')

        # Keep the run genome explicit in outputs for auditability/reproducibility.
        

        self.annotated_df = annotated_df
        if utils.out_path:
            self.write_output()

    def preprocess_fusion_bkpt(self):
        logger.info("-- Pre-processing input fusion breakpoints")
        prep_obj = PrepareInputBkpt(self.fusion_bkpt, self.format,
                                    self.target_df, utils.out_path)
        filtered_bkpts, excluded_bkpts, total_count = prep_obj.preprocess_bkpt()
        logger.info(f'Excluded breakpoints: {excluded_bkpts.shape[0]}')

        if utils.out_path:
            exbkpt_file = os.path.join(utils.out_path, 'excluded_breakpoints.txt')
            logger.info(f'Written excluded breakpoint to {exbkpt_file}')
            excluded_bkpts.to_csv(exbkpt_file, index=False, sep='\t')

        return filtered_bkpts

    def add_exon_annotation(self, fusion_bkpt):
        return ex_annot.annotate_fusion_exons(fusion_bkpt)

    def add_cosmic_gtex_annots(self, annotated_df):
        return cos_gtex.annotate_gtex_cosmic(annotated_df)

    def add_frame_annotation(self, annotated_df_gtex_cos):
        return frame_calc.frame_calculation(annotated_df_gtex_cos)

    def add_mane_annotation(self, annotated_df):
        return mane_annot.add_mane(annotated_df)

    def add_fusion_sequence(self, annotated_df_frame_calc):
        return fus_seq.fusion_sequence(annotated_df_frame_calc, self.ref)

    def add_black_list_genes(self, annotated_df_bl_list):
        return bl_list.annotate_bl_list(annotated_df_bl_list)

    def join_op(self, file):
        return os.path.join(utils.out_path, file)

    @staticmethod
    def _display_genome(genome):
        if not genome:
            return 'GRCh37'
        mapping = {
            'grch37': 'GRCh37',
            'grch38': 'GRCh38',
        }
        return mapping.get(str(genome).lower(), str(genome))

    def write_output(self):
        logger.info("-- Writing final output")
        html = self.join_op('pyfuse_fusion_annotation.html')
        fexcel = self.join_op('pyfuse_output.xlsx')
        summary = self.join_op('pyfuse_fusion_summary.txt')
        fvcf = self.join_op('pyfuse_output.vcf')

        #pivot_ui(self.annotated_df, outfile_path=html)
        self.annotated_df = self.annotated_df.apply(
                                lambda col: col.map(lambda x: x.replace('"', '') if isinstance(x, str) else x)
                            )

        fusion_counts = self.annotated_df[['Fusion_Position', "5'-3'Gene_Partners", 'Frame_Status']].value_counts().reset_index(name='count')
        fusion_counts.to_csv(summary, sep='\t', index=False)

        # Reorder columns: move Frame_Status to position 5 (index 4)
        output_df = self.annotated_df.copy()
        cols = list(output_df.columns)
        if 'Frame_Status' in cols:
            cols.remove('Frame_Status')
            cols.insert(4, 'Frame_Status')
            output_df = output_df[cols]

        # Write Excel with auto-expanded columns
        output_df.to_excel(fexcel, index=False)
        self._auto_expand_xlsx_columns(fexcel)

        # Write BND-style VCF output
        try:
            fusion_df_to_bnd_vcf(self.annotated_df, fvcf)
            
        except Exception as e:
            logger.error(f'Failed to write VCF: {e}')

        final_df = output_df.copy()
        final_df, embedded_fusion_plots = add_fusion_plot_links(
            final_df,
            utils.out_path,
            mode=self.fusion_plot_mode,
        )
        final_df["5'-3'Gene_Partners"] = final_df["5'-3'Gene_Partners"].apply(lambda x: utils.gene_pair_to_html_links(x))
        for coord_col in ["5'co-ordinate", "3'co-ordinate"]:
            if coord_col in final_df.columns:
                final_df[coord_col] = final_df[coord_col].apply(lambda x: utils.coord_to_ucsc_link(x))
        for exon_col in ["5'_Exon_Annotation", "3'_Exon_Annotation"]:
            if exon_col in final_df.columns:
                final_df[exon_col] = final_df[exon_col].apply(lambda x: utils.exon_annotation_to_refseq_link(x))

        write_df2html(
            final_df,
            html,
            title="PyFuse Gene Fusion Annotations",
            embedded_fusion_plots=embedded_fusion_plots,
        )


        logger.info(f'output1 -- {html}')
        logger.info(f'output2 -- {fexcel}')
        logger.info(f'output3 -- {fvcf}')

    def _auto_expand_xlsx_columns(self, fexcel):
        """Auto-expand columns in XLSX file based on content width."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(fexcel)
            ws = wb.active
            
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(fexcel)
        except Exception as e:
            logger.warning(f'Failed to auto-expand XLSX columns: {e}')
